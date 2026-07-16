import os
from openpyxl import load_workbook
from .base import log

def convert_xlsx_to_md(xlsx_path, log_fn=None):

    if xlsx_path.endswith('.xls'):
        log("✗ 不支持 .xls 格式，请使用 .xlsx 格式", log_fn)
        return None

    md_path = xlsx_path.replace('.xlsx', '.md')
    try:
        wb = load_workbook(xlsx_path)
        body = f"# {os.path.basename(xlsx_path)}\n\n"

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            body += f"## {sheet_name}\n\n"

            max_row = ws.max_row
            max_col = ws.max_column

            if max_row == 0 or max_col == 0:
                body += "(空工作表)\n\n"
                continue

            headers = []
            for col in range(1, max_col + 1):
                cell = ws.cell(row=1, column=col)
                headers.append(str(cell.value) if cell.value is not None else '')

            md_table = "| " + " | ".join(headers) + " |\n"
            md_table += "| " + " | ".join("---" for _ in headers) + " |\n"

            for row in range(2, max_row + 1):
                row_data = []
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    row_data.append(str(cell.value) if cell.value is not None else '')
                md_table += "| " + " | ".join(row_data) + " |\n"

            body += md_table + "\n"

        with open(md_path, 'w', encoding='utf-8') as f:
            f.write(body)

        log(f"✓ 转换成功: {xlsx_path} -> {md_path}", log_fn)
        return md_path, body
    except Exception as e:
        log(f"✗ 转换失败: {xlsx_path}", log_fn)
        log(f"  错误: {str(e)}", log_fn)
        return None